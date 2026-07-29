#!/usr/bin/env python3
"""Parse Tank1..Tank4 calibration sheets from the CAPTAIN VENIAMIS workbook -> JSON."""
from __future__ import annotations
import json, sys, re
from openpyxl import load_workbook

def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)

def read_number_row(ws, r, start_c, max_c=40):
    vals = []
    c = start_c
    while c <= max_c:
        v = ws.cell(r, c).value
        if not is_num(v):
            break
        vals.append(v)
        c += 1
    return vals

def guess_meta(name: str):
    n = name.upper()
    fuel_role = 'other'
    if re.search(r'SETT', n): fuel_role = 'settling'
    elif re.search(r'SERVICE', n): fuel_role = 'service'
    elif re.search(r'OVERFLOW', n): fuel_role = 'overflow'
    elif re.search(r'STORAGE|STOR\.|H\.?F\.?O\.?\s*TANK|MDO|MGO|NO\.\d', n): fuel_role = 'storage'

    side = 'center'
    if re.search(r'\(P\)|\.P\b|PORT|TK\.P', n): side = 'port'
    elif re.search(r'\(S\)|\.S\b|STBD|STARBOARD|TK\.S', n): side = 'starboard'

    tank_no = None
    m = re.search(r'NO\.?\s*(\d+)', n) or re.search(r'\((\d+)\)', n)
    if m: tank_no = int(m.group(1))

    fuel_grade = 'other'
    if re.search(r'LS\s*H\.?F\.?O|VLSFO|LSFO', n): fuel_grade = 'lsfo'
    elif re.search(r'H\.?F\.?O', n): fuel_grade = 'hfo'
    elif 'MDO' in n: fuel_grade = 'mdo'
    elif 'MGO' in n: fuel_grade = 'mgo'
    return dict(fuelRole=fuel_role, side=side, tankNo=tank_no, fuelGrade=fuel_grade)


def guess_category(name: str) -> str:
    nu = name.upper()
    if re.search(r'L\.?O\.|LUBE|CYL|SUMP', nu) and not re.search(r'H\.?F\.?O|MDO|MGO|F\.?O\b', nu):
        return 'lube'
    if re.search(r'F\.?W\.|WATER|DISTILLED|DRINKING', nu):
        return 'water'
    if re.search(r'BILGE|SLUDGE|SEWAGE|DRAIN|STERN\s*TUBE|OVERFLOW', nu) and not re.search(
        r'H\.?F\.?O|MDO|MGO|NO\.\d', nu
    ):
        return 'misc'
    if re.search(r'H\.?F\.?O|MDO|MGO|F\.?O\.?\s|FUEL', nu) or re.search(r'NO\.\d', nu):
        return 'fuel'
    if re.search(r'L\.?O\.|LUBE|CYL|SUMP', nu):
        return 'lube'
    return 'misc'


def _expand_abbrev(s: str) -> str:
    s = str(s).upper()
    # Insert a separator so H.F.O.SETTLING → HFO SETTLING (not HFOSETTLING)
    s = re.sub(r'H\.?\s*F\.?\s*O\.?(?=[A-Z])', 'HFO ', s)
    s = re.sub(r'H\.?\s*F\.?\s*O\.?', 'HFO', s)
    s = re.sub(r'M\.?\s*D\.?\s*O\.?(?=[A-Z])', 'MDO ', s)
    s = re.sub(r'M\.?\s*D\.?\s*O\.?', 'MDO', s)
    s = re.sub(r'M\.?\s*G\.?\s*O\.?(?=[A-Z])', 'MGO ', s)
    s = re.sub(r'M\.?\s*G\.?\s*O\.?', 'MGO', s)
    s = re.sub(r'L\.?\s*O\.?(?=[A-Z])', 'LO ', s)
    s = re.sub(r'L\.?\s*O\.?', 'LO', s)
    s = re.sub(r'F\.?\s*W\.?(?=[A-Z])', 'FW ', s)
    s = re.sub(r'F\.?\s*W\.?', 'FW', s)
    s = re.sub(r'\bLS\s*HFO\b|\bLSHFO\b', 'LSHFO', s)
    return s


def norm_name(s: str) -> str:
    return re.sub(r'[^A-Z0-9]', '', _expand_abbrev(s))


def name_tokens(s: str):
    return set(re.findall(r'[A-Z0-9]+', _expand_abbrev(s)))


def setup_match_score(tank_name: str, setup_name: str) -> int:
    """Fuzzy score for Setup ↔ tank title (Giorgis uses shorter Setup labels)."""
    a, b = norm_name(tank_name), norm_name(setup_name)
    if not a or not b:
        return 0
    if a == b:
        return 1000
    if a in b or b in a:
        return 800 + min(len(a), len(b))
    ta, tb = name_tokens(tank_name), name_tokens(setup_name)
    ta2 = {(t if t not in ('STOR',) else 'STORAGE') for t in ta} - {'TK', 'TANK', 'THE', 'AND'}
    tb2 = {(t if t not in ('STOR',) else 'STORAGE') for t in tb} - {'TK', 'TANK', 'THE', 'AND'}
    if not ta2 or not tb2:
        return 0
    inter = ta2 & tb2
    if not inter:
        return 0
    # Side-only (P/S) or role-only matches are too weak across tank families
    weak = {'P', 'S', 'PORT', 'STBD', 'STARBOARD', 'STORAGE', 'SERVICE', 'SETTLING', 'OVERFLOW', 'NO'}
    strong = inter - weak
    if not strong:
        return 0
    # Conflicting fuel grades → reject
    grades = {'HFO', 'MDO', 'MGO', 'LSHFO', 'FW', 'LO'}
    ga, gb = ta2 & grades, tb2 & grades
    if ga and gb and ga != gb:
        return 0
    return int(100 * len(inter) / max(len(ta2), len(tb2))) + 20 * len(strong) + 5 * len(inter)


def even_keel_index(trim_vals):
    if not trim_vals:
        return None
    best_i, best_abs = 0, abs(trim_vals[0])
    for i, v in enumerate(trim_vals):
        if abs(v) < best_abs:
            best_i, best_abs = i, abs(v)
    return best_i


def robust_grid_capacity(trim_vals, trim_grid):
    """
    Capacity from a volume grid, ignoring OCR/Excel outliers.
    Prefer the even-keel column; fall back to row-0 / column medians.
    """
    if not trim_grid or not trim_grid[0]:
        return 0
    ek = even_keel_index(trim_vals)
    cols = []
    if ek is not None:
        cols = [row[ek] for row in trim_grid if ek < len(row) and is_num(row[ek])]
    if len(cols) < 3:
        cols = [v for row in trim_grid for v in row if is_num(v)]
    pos = sorted(v for v in cols if v > 0)
    if not pos:
        return 0
    # Reject values far above the high percentile (comma/decimal OCR blow-ups)
    p90 = pos[int(0.9 * (len(pos) - 1))]
    clean = [v for v in pos if v <= max(p90 * 3, p90 + 50)]
    return max(clean) if clean else pos[-1]


def looks_like_volume_grid(trim_grid, volume_curve=None, trim_axis=None):
    """
    Giorgis Tank1: trim block holds CAPACITY m³ (often flat ~full at ullage 0).
    Veniamis Tank1: trim block holds small corrections + separate volume curve.
    """
    if not trim_grid or not trim_grid[0]:
        return False
    first = [v for v in trim_grid[0] if is_num(v)]
    flat = [v for row in trim_grid for v in row if is_num(v)]
    if len(flat) < 8 or not first:
        return False
    first_abs = [abs(v) for v in first]
    first_med = sorted(first_abs)[len(first_abs) // 2]
    first_span = max(first) - min(first)
    vmax = max(abs(v) for v in flat)

    vc_vals = [v for v in (volume_curve or {}).get('v') or [] if is_num(v)]
    vc_max = max((abs(v) for v in vc_vals), default=0)

    # Explicit volume curve labeled and larger than trim corrections → Veniamis-style
    if len(vc_vals) >= 20 and vc_max >= 20 and vc_max >= vmax * 0.8 and min(vc_vals) >= -1:
        return False

    # Nearly-full first row (ullage/sounding 0) with large flat volumes → Giorgis capacity grid
    if first_med >= 20 and first_span <= max(first_med * 0.2, 5):
        return True

    # Monotonic growth along axis → sounding-from-bottom volume table
    if trim_axis and len(trim_axis) >= 6:
        ek_vals = []
        for row in trim_grid:
            if not row:
                continue
            idx = 1 if len(row) > 1 else 0
            if is_num(row[idx]):
                ek_vals.append(row[idx])
        if len(ek_vals) >= 6:
            head = sum(ek_vals[:3]) / 3
            tail = sum(ek_vals[-3:]) / 3
            if tail > head * 1.5 and tail >= 5 and vc_max < tail * 0.5:
                return True

    # Large grid without a real volume curve
    if vmax >= 40 and vc_max < vmax * 0.5:
        return True
    return False


def sanitize_volume_grid(trim_grid, capacity_hint=None):
    """Clamp obvious OCR/Excel blow-ups in volume grids (e.g. 463476 vs 463.476)."""
    if not trim_grid:
        return trim_grid
    flat = [v for row in trim_grid for v in row if is_num(v) and v > 0]
    if len(flat) < 5:
        return trim_grid
    flat_sorted = sorted(flat)
    p90 = flat_sorted[int(0.9 * (len(flat_sorted) - 1))]
    limit = max(p90 * 3, (capacity_hint or 0) * 1.5, p90 + 50)
    if limit <= 0:
        return trim_grid
    out = []
    for row in trim_grid:
        out.append([
            (0 if (is_num(v) and v > limit) else v) if is_num(v) else 0
            for v in row
        ])
    return out


def promote_volume_grid(tank: dict) -> dict:
    """Convert a mis-tagged correction parse into a direct ullage/sounding volume grid."""
    cap = robust_grid_capacity(tank.get('trimVals') or [], tank.get('trimGrid') or [])
    tank['trimGrid'] = sanitize_volume_grid(tank.get('trimGrid') or [], cap)
    cap = robust_grid_capacity(tank.get('trimVals') or [], tank.get('trimGrid') or []) or cap
    tank['calcType'] = 'direct'
    tank['correctionDivisor'] = 1
    tank['volumeCurve'] = {'x': [], 'v': []}
    if cap > 0:
        tank['capacity'] = cap
    axis = tank.get('trimAxis') or []
    grid = tank.get('trimGrid') or []
    if len(axis) >= 6 and grid:
        idx = 1 if len(grid[0]) > 1 else 0
        head = [row[idx] for row in grid[:3] if row and is_num(row[idx])]
        tail = [row[idx] for row in grid[-3:] if row and is_num(row[idx])]
        if head and tail and (sum(tail) / len(tail)) > (sum(head) / len(head)) * 1.5:
            tank['soundingMethod'] = 'sounding'
    return tank

def parse_correction(ws, title_row):
    name = str(ws.cell(title_row, 1).value or '').strip()
    header = title_row + 3
    trim_vals = read_number_row(ws, header, 2)
    if len(trim_vals) < 3:
        return None
    list_start = None
    for c in range(2 + len(trim_vals), 45):
        if is_num(ws.cell(header, c).value):
            list_start = c
            break
    list_vals = read_number_row(ws, header, list_start) if list_start else []

    trim_axis, trim_grid = [], []
    list_axis, list_grid = [], []
    vol_map = {}  # x -> v
    r = header + 1
    while True:
        sounding = ws.cell(r, 1).value
        if not is_num(sounding):
            break
        trim_axis.append(sounding)
        trim_grid.append([ws.cell(r, 2 + i).value if is_num(ws.cell(r, 2 + i).value) else 0 for i in range(len(trim_vals))])
        if list_start and list_vals:
            la = ws.cell(r, list_start - 1).value
            if is_num(la):
                list_axis.append(la)
                list_grid.append([ws.cell(r, list_start + i).value if is_num(ws.cell(r, list_start + i).value) else 0 for i in range(len(list_vals))])
        r += 1
        if r > header + 5000:
            break

    # Volume curve: Veniamis labels "SOUNDING CM" / "sounding VOLUME" on title_row+1.
    # Giorgis has heel headers in those columns — never treat heel as volume.
    vol_map = {}
    vol_x_col = vol_v_col = None
    for lr in (title_row + 1, title_row + 2):
        for c in range(2, 30):
            lab = ws.cell(lr, c).value
            if not isinstance(lab, str):
                continue
            low = re.sub(r'\s+', ' ', lab.strip().lower())
            if re.search(r'sounding\s*cm|depth\s*cm', low):
                nxt = str(ws.cell(lr, c + 1).value or '')
                nxt_low = nxt.lower()
                if re.search(r'volume|vol\.?\b|m3|m³|capacity', nxt_low):
                    vol_x_col, vol_v_col = c, c + 1
                    break
                # Giorgis reference pair sounding cm | sounding ullage — skip
                if 'ullage' in nxt_low:
                    continue
            if re.search(r'^sounding\s*volume$|volume\s*m', low) and vol_v_col is None:
                vol_v_col = c
                vol_x_col = c - 1
        if vol_x_col and vol_v_col:
            break

    if vol_x_col and vol_v_col and vol_v_col != vol_x_col:
        for rr in range(header + 1, header + 5000):
            vx, vv = ws.cell(rr, vol_x_col).value, ws.cell(rr, vol_v_col).value
            if is_num(vx) and is_num(vv) and vv >= -0.01:
                vol_map[vx] = vv
            elif rr > header + max(len(trim_axis), 50) and not is_num(vx) and not is_num(vv):
                if rr > header + len(trim_axis) + 200 and not vol_map:
                    break
                if rr > header + max(len(vol_map), len(trim_axis)) + 50:
                    break

    # Fallback: Veniamis often keeps sounding/volume in columns 12/13 even when labels drift.
    # Only when the trim block looks like corrections (not a Giorgis capacity grid).
    if not vol_map and trim_grid:
        first = [v for v in trim_grid[0] if is_num(v)]
        first_med = sorted(abs(v) for v in first)[len(first) // 2] if first else 0
        if first_med < 40:
            probe = []
            for rr in range(header + 1, header + 12):
                vx, vv = ws.cell(rr, 12).value, ws.cell(rr, 13).value
                if is_num(vx) and is_num(vv) and vv >= -0.01 and vx >= 0:
                    probe.append((vx, vv))
            xs_p = [x for x, _ in probe]
            # Real sounding axis starts near 0 and increases; heel columns do not.
            if (
                len(probe) >= 4
                and xs_p[0] <= 20
                and xs_p == sorted(xs_p)
                and max(xs_p) >= 50
            ):
                for rr in range(header + 1, header + 5000):
                    vx, vv = ws.cell(rr, 12).value, ws.cell(rr, 13).value
                    if is_num(vx) and is_num(vv) and vv >= -0.01:
                        vol_map[vx] = vv
                    elif rr > header + max(len(trim_axis), 50) and not is_num(vx) and not is_num(vv):
                        if rr > header + max(len(vol_map), len(trim_axis)) + 50:
                            break

    if not trim_axis:
        return None
    xs = sorted(vol_map)
    vs = [vol_map[x] for x in xs]
    capacity = max(vs) if vs else 0
    meta = guess_meta(name)

    def detect_inc(axis):
        if len(axis) < 2:
            return 1
        diffs = [round(abs(axis[i] - axis[i - 1]) * 1000) / 1000 for i in range(1, len(axis)) if axis[i] != axis[i - 1]]
        if not diffs:
            return 1
        from collections import Counter
        best = Counter(diffs).most_common(1)[0][0]
        for p in (1, 2, 5, 10, 20, 25, 50):
            if abs(best - p) < 1e-6:
                return p
        return best

    tank = {
        'name': name,
        'category': guess_category(name),
        'calcType': 'correction',
        'correctionDivisor': 10,
        'soundingMethod': 'ullage',
        'soundingIncrement': detect_inc(trim_axis),
        'heelIncrement': detect_inc(list_axis) if list_axis else detect_inc(trim_axis),
        'trimAxis': trim_axis,
        'trimVals': trim_vals,
        'trimGrid': trim_grid,
        'listAxis': list_axis,
        'listVals': list_vals,
        'listGrid': list_grid,
        'volumeCurve': {'x': xs, 'v': vs},
        'capacity': capacity,
        **meta,
    }
    if looks_like_volume_grid(trim_grid, tank['volumeCurve'], trim_axis):
        promote_volume_grid(tank)
    return tank

def is_axis_header(v):
    if not isinstance(v, str):
        return False
    s = re.sub(r'\s+', ' ', v.strip().lower())
    return s in ('depth', 'gauge', 'gauge ull', 'gauge ullage', 'ullage', 'sounding')

def parse_direct(ws, title_row):
    raw = str(ws.cell(title_row, 1).value or '').strip()
    name = re.sub(r'\s*-\s*Volume in m3\s*$', '', raw, flags=re.I).strip()
    header = title_row + 1
    depth = ws.cell(header, 1).value
    if not is_axis_header(depth):
        return None
    trim_vals = read_number_row(ws, header, 2)
    if len(trim_vals) < 2:
        return None

    heel_depth_col = heel_start = None
    for c in range(2 + len(trim_vals), 45):
        v = ws.cell(header, c).value
        if is_axis_header(v) or (isinstance(v, str) and v.strip().lower() == 'ullage'):
            # heel/list table may start at next numeric col (Ullage label then Depth then values)
            heel_depth_col = c
            for cc in range(c, c + 20):
                if is_num(ws.cell(header, cc).value):
                    heel_start = cc
                    # if label itself wasn't depth axis, axis values are in column c when numeric rows use col c
                    if not is_num(ws.cell(header + 1, c).value):
                        heel_depth_col = cc - 1 if cc > c else c
                    break
            break
    # Tank2 style: "Ullage" then "Depth" then heel headers
    if heel_start is None:
        for c in range(2 + len(trim_vals), 45):
            v = ws.cell(header, c).value
            if isinstance(v, str) and 'ullage' in v.strip().lower():
                for cc in range(c + 1, c + 15):
                    if is_axis_header(ws.cell(header, cc).value):
                        heel_depth_col = cc
                    if is_num(ws.cell(header, cc).value):
                        heel_start = cc
                        if heel_depth_col is None:
                            heel_depth_col = cc - 1
                        break
                break
    list_vals = read_number_row(ws, header, heel_start) if heel_start else []

    trim_axis, trim_grid = [], []
    list_axis, list_grid = [], []
    r = header + 1
    while True:
        d = ws.cell(r, 1).value
        if not is_num(d):
            break
        trim_axis.append(d)
        trim_grid.append([ws.cell(r, 2 + i).value if is_num(ws.cell(r, 2 + i).value) else 0 for i in range(len(trim_vals))])
        if heel_depth_col and heel_start and list_vals:
            la = ws.cell(r, heel_depth_col).value
            if is_num(la):
                list_axis.append(la)
                list_grid.append([ws.cell(r, heel_start + i).value if is_num(ws.cell(r, heel_start + i).value) else 0 for i in range(len(list_vals))])
        r += 1
        if r > header + 5000:
            break
    if not trim_axis:
        return None

    capacity = robust_grid_capacity(trim_vals, trim_grid)
    trim_grid = sanitize_volume_grid(trim_grid, capacity)
    capacity = robust_grid_capacity(trim_vals, trim_grid) or capacity
    return {
        'name': name,
        'category': guess_category(name),
        'calcType': 'direct',
        'correctionDivisor': 1,
        'soundingMethod': 'sounding',
        'trimAxis': trim_axis,
        'trimVals': trim_vals,
        'trimGrid': trim_grid,
        'listAxis': list_axis,
        'listVals': list_vals,
        'listGrid': list_grid,
        'volumeCurve': {'x': [], 'v': []},
        'capacity': capacity,
        **guess_meta(name),
    }

def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'TANK MANAGEMENT CAPTAIN VENIAMIS FINAL VERSION.xlsm'
    wb = load_workbook(path, data_only=False, keep_vba=False)
    tanks = {'fuel': [], 'lube': [], 'misc': [], 'water': []}
    found = []

    for sheet_name in wb.sheetnames:
        if not re.match(r'Tank\d+', sheet_name, re.I):
            continue
        ws = wb[sheet_name]
        # limit scan using max_row but Tank1 is huge — scan col A only
        max_r = min(ws.max_row or 0, 12000)
        for r in range(1, max_r + 1):
            a = ws.cell(r, 1).value
            if not isinstance(a, str):
                continue
            s = a.strip()
            if not s or s == '`' or re.match(r'(?i)sounding|depth$', s):
                continue
            tank = None
            next_a = ws.cell(r + 1, 1).value
            if is_axis_header(next_a):
                tank = parse_direct(ws, r)
            elif re.search(r'TANK|TK\.|H\.?F\.?O|MDO|MGO|WATER|BILGE|SLUDGE|SEWAGE|DRAIN|STERN', s, re.I):
                # Prefer direct when row+1 looks like a gauge header row with numeric trim vals
                if is_num(ws.cell(r + 1, 2).value) and isinstance(next_a, str):
                    tank = parse_direct(ws, r)
                if not tank:
                    hdr = read_number_row(ws, r + 3, 2)
                    if len(hdr) >= 3:
                        tank = parse_correction(ws, r)
            if not tank:
                continue
            cat = tank['category']
            if cat not in tanks:
                tanks[cat] = []
            if any(t['name'] == tank['name'] for t in tanks[cat]):
                continue
            tank['id'] = f"{cat}{len(tanks[cat]) + 1}"
            tanks[cat].append(tank)
            found.append({
                'sheet': sheet_name,
                'row': r,
                'name': tank['name'],
                'calcType': tank['calcType'],
                'rows': len(tank['trimAxis']),
                'capacity': tank.get('capacity'),
            })

    # Setup pipe / capacity overlay
    setup = []
    if 'Setup' in wb.sheetnames:
        ws = wb['Setup']
        for r in range(2, 80):
            name = ws.cell(r, 1).value
            if not isinstance(name, str) or not name.strip():
                continue
            if name.strip().startswith('='):
                continue
            setup.append({
                'name': name.strip(),
                'pipeHeight': ws.cell(r, 3).value if is_num(ws.cell(r, 3).value) else ws.cell(r, 2).value,
                'capacity100': ws.cell(r, 9).value,
            })
        for cat, arr in tanks.items():
            for t in arr:
                best, best_score = None, 0
                for s in setup:
                    score = setup_match_score(t['name'], s['name'])
                    if score > best_score:
                        best, best_score = s, score
                if not best or best_score < 40:
                    continue
                if is_num(best.get('pipeHeight')):
                    t['pipeHeight'] = best['pipeHeight']
                if is_num(best.get('capacity100')):
                    # Setup 100% capacity is authoritative when present
                    t['capacity'] = best['capacity100']
                t['setupMatch'] = best['name']

    json.dump({'tanks': tanks, 'found': found, 'setup': setup}, sys.stdout)

if __name__ == '__main__':
    main()
