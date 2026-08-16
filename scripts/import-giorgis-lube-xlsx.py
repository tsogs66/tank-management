#!/usr/bin/env python3
"""Extract all Giorgis lube-oil tank calibration blocks from an XLSX workbook."""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install -r requirements.txt"}))
    sys.exit(1)


def to_num(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    try:
        return float(str(value).strip().replace("\u2212", "-").replace(",", ""))
    except ValueError:
        return None


def clean_num(value):
    number = to_num(value)
    if number is None:
        return None
    return int(number) if number.is_integer() else number


def detect_increment(axis):
    diffs = [
        round(abs(axis[i] - axis[i - 1]), 6)
        for i in range(1, len(axis))
        if axis[i] != axis[i - 1]
    ]
    return Counter(diffs).most_common(1)[0][0] if diffs else 1


def clean_title(value):
    title = str(value or "").strip()
    title = re.sub(r"\s*-\s*Volume\s+in\s+m(?:3|³)\s*$", "", title, flags=re.I)
    title = re.sub(r"^\d+\s+compart:\s*", "", title, flags=re.I)
    return re.sub(r"\s+", " ", title).strip(" .")


def tank_role(title):
    upper = title.upper()
    if "SETT" in upper:
        return "settling"
    if "SUMP" in upper:
        return "other"
    return "storage"


def side_from_title(title):
    upper = title.upper()
    if re.search(r"\(\s*P\s*\)|\bPORT\b", upper):
        return "port"
    if re.search(r"\(\s*S\s*\)|\bSTBD\b|\bSTARBOARD\b", upper):
        return "starboard"
    return "center"


def numeric_header(row, start, stop):
    values = []
    for col in range(start, stop + 1):
        number = clean_num(row[col - 1] if col - 1 < len(row) else None)
        if number is None:
            if values:
                break
            continue
        values.append(number)
    return values


def robust_capacity(trim_vals, trim_grid):
    if not trim_grid:
        return 0
    try:
        col = trim_vals.index(0)
    except ValueError:
        col = min(range(len(trim_vals)), key=lambda i: abs(trim_vals[i]))
    values = [row[col] for row in trim_grid if col < len(row) and row[col] >= 0]
    return round(max(values), 6) if values else 0


def parse_block(rows, header_index, title):
    header = rows[header_index]
    # Giorgis layout: A=depth, B:I=trim volumes, J=gap, K=heel depth, L:T=heel correction.
    trim_vals = numeric_header(header, 2, 10)
    list_vals = numeric_header(header, 12, len(header))
    if len(trim_vals) < 2:
        return None

    trim_axis, trim_grid = [], []
    list_axis, list_grid = [], []
    for row in rows[header_index + 1:]:
        depth = clean_num(row[0] if row else None)
        if depth is None:
            break

        trim_row = []
        valid_trim = 0
        for j in range(len(trim_vals)):
            number = clean_num(row[j + 1] if j + 1 < len(row) else None)
            if number is not None:
                valid_trim += 1
            trim_row.append(number if number is not None else 0)
        if valid_trim:
            trim_axis.append(depth)
            trim_grid.append(trim_row)

        heel_depth = clean_num(row[10] if len(row) > 10 else None)
        if heel_depth is not None and list_vals:
            list_row = []
            valid_list = 0
            for j in range(len(list_vals)):
                number = clean_num(row[j + 11] if j + 11 < len(row) else None)
                if number is not None:
                    valid_list += 1
                list_row.append(number if number is not None else 0)
            if valid_list:
                list_axis.append(heel_depth)
                list_grid.append(list_row)

    if len(trim_axis) < 2:
        return None

    try:
        even_col = trim_vals.index(0)
    except ValueError:
        even_col = min(range(len(trim_vals)), key=lambda i: abs(trim_vals[i]))
    volume_curve = {
        "x": list(trim_axis),
        "v": [row[even_col] for row in trim_grid],
    }
    has_list_signal = any(abs(value) > 1e-12 for row in list_grid for value in row)

    return {
        "name": clean_title(title),
        "category": "lube",
        "fuelRole": tank_role(title),
        "fuelGrade": "other",
        "side": side_from_title(title),
        "tankNo": None,
        "calcType": "direct",
        "capacity": robust_capacity(trim_vals, trim_grid),
        "pipeHeight": 0,
        "soundingMethod": "sounding",
        "correctionDivisor": 1,
        "soundingIncrement": detect_increment(trim_axis),
        "heelIncrement": detect_increment(list_axis) if len(list_axis) > 1 else detect_increment(trim_axis),
        "trimAxis": trim_axis,
        "trimVals": trim_vals,
        "trimGrid": trim_grid,
        "listAxis": list_axis if has_list_signal else [],
        "listVals": list_vals if has_list_signal else [],
        "listGrid": list_grid if has_list_signal else [],
        "volumeCurve": volume_curve,
        "importFormat": "giorgis-lube-xlsx",
        "pdfSource": clean_title(title),
    }


def previous_title(rows, header_index):
    for index in range(header_index - 1, max(-1, header_index - 12), -1):
        first = rows[index][0] if rows[index] else None
        if not isinstance(first, str) or not first.strip():
            continue
        upper = first.upper()
        if any(token in upper for token in ("VOLUME IN", "L.O.", "L O ", "CYL.O", "TK.")):
            return first
    return ""


def extract(path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    tanks = []
    warnings = []
    sheets = []
    for sheet in workbook.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        sheet_count = 0
        for index, row in enumerate(rows):
            first = str(row[0] or "").strip().upper() if row else ""
            if first not in ("DEPTH", "SOUNDING", "ULLAGE"):
                continue
            title = previous_title(rows, index)
            if not title:
                warnings.append(f"{sheet.title} row {index + 1}: table has no tank title")
                continue
            tank = parse_block(rows, index, title)
            if tank:
                tanks.append(tank)
                sheet_count += 1
            else:
                warnings.append(f"{clean_title(title)}: no usable calibration rows")
        sheets.append({"name": sheet.title, "tankCount": sheet_count})

    if not tanks:
        raise ValueError("No Giorgis lube tank tables found")
    return {
        "format": "giorgis-lube-xlsx",
        "tankCount": len(tanks),
        "sheets": sheets,
        "warnings": warnings,
        "tanks": tanks,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook")
    args = parser.parse_args()
    try:
        json.dump(extract(args.workbook), sys.stdout)
    except Exception as exc:
        json.dump({"error": str(exc), "tanks": []}, sys.stdout)
        sys.exit(1)


if __name__ == "__main__":
    main()
