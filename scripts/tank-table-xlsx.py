#!/usr/bin/env python3
"""
Export / import a single tank calibration table as .xlsx for editing in Excel.

Usage:
  # Export tank JSON -> xlsx
  python3 scripts/tank-table-xlsx.py export --in tank.json --out tank.xlsx

  # Import xlsx -> calibration JSON patch
  python3 scripts/tank-table-xlsx.py import --in tank.xlsx

Sheets: META, TRIM, VOLUME, LIST
"""
from __future__ import annotations

import argparse
import json
import sys
from collections import Counter

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}))
    sys.exit(1)


META_KEYS = [
    "calcType",
    "capacity",
    "correctionDivisor",
    "pipeHeight",
    "soundingMethod",
    "soundingIncrement",
    "heelIncrement",
    "name",
]


def is_num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def to_num(v):
    if v is None or v == "":
        return None
    if is_num(v):
        return float(v)
    s = str(v).strip().replace("\u2212", "-").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def detect_inc(axis):
    if len(axis) < 2:
        return 1
    diffs = [
        round(abs(axis[i] - axis[i - 1]) * 1000) / 1000
        for i in range(1, len(axis))
        if axis[i] != axis[i - 1]
    ]
    if not diffs:
        return 1
    best = Counter(diffs).most_common(1)[0][0]
    for p in (1, 2, 5, 10, 20, 25, 50):
        if abs(best - p) < 1e-6:
            return p
    return best


def export_xlsx(tank: dict, out_path: str):
    wb = Workbook()

    ws = wb.active
    ws.title = "META"
    ws.append(["key", "value"])
    ws.append(["id", tank.get("id", "")])
    for k in META_KEYS:
        if tank.get(k) is None or tank.get(k) == "":
            continue
        ws.append([k, tank.get(k)])

    trim_vals = tank.get("trimVals") or []
    trim_axis = tank.get("trimAxis") or []
    trim_grid = tank.get("trimGrid") or []
    ws_t = wb.create_sheet("TRIM")
    ws_t.append(["SOUNDING", *trim_vals])
    for i, s in enumerate(trim_axis):
        row = trim_grid[i] if i < len(trim_grid) else []
        ws_t.append([s, *[row[j] if j < len(row) and row[j] is not None else 0 for j in range(len(trim_vals))]])

    vx = (tank.get("volumeCurve") or {}).get("x") or []
    vv = (tank.get("volumeCurve") or {}).get("v") or []
    ws_v = wb.create_sheet("VOLUME")
    ws_v.append(["SOUNDING", "VOLUME"])
    for i, x in enumerate(vx):
        ws_v.append([x, vv[i] if i < len(vv) else 0])

    list_vals = tank.get("listVals") or []
    list_axis = tank.get("listAxis") or []
    list_grid = tank.get("listGrid") or []
    ws_l = wb.create_sheet("LIST")
    ws_l.append(["SOUNDING", *list_vals])
    for i, s in enumerate(list_axis):
        row = list_grid[i] if i < len(list_grid) else []
        ws_l.append([s, *[row[j] if j < len(row) and row[j] is not None else 0 for j in range(len(list_vals))]])

    # Also a combined "TABLE" sheet matching the app grid for casual editing
    ws_c = wb.create_sheet("TABLE", 0)
    ws_c.append(["Tank", tank.get("name", ""), "id", tank.get("id", "")])
    ws_c.append([])
    header = ["SOUNDING", *trim_vals]
    if vx:
        header += ["SOUNDING_CM", "VOLUME"]
    if list_vals:
        header += ["LIST_SOUNDING", *list_vals]
    ws_c.append(header)
    n = max(len(trim_axis), len(list_axis), len(vx), 0)
    vol_map = {float(x): vv[i] if i < len(vv) else 0 for i, x in enumerate(vx)}
    for i in range(n):
        row = []
        if i < len(trim_axis):
            row.append(trim_axis[i])
            g = trim_grid[i] if i < len(trim_grid) else []
            row.extend([g[j] if j < len(g) and g[j] is not None else 0 for j in range(len(trim_vals))])
        else:
            row.extend([None] * (1 + len(trim_vals)))
        if vx:
            if i < len(trim_axis):
                row.append(trim_axis[i])
                row.append(vol_map.get(float(trim_axis[i]), None))
            elif i < len(vx):
                row.append(vx[i])
                row.append(vv[i] if i < len(vv) else 0)
            else:
                row.extend([None, None])
        if list_vals:
            if i < len(list_axis):
                row.append(list_axis[i])
                g = list_grid[i] if i < len(list_grid) else []
                row.extend([g[j] if j < len(g) and g[j] is not None else 0 for j in range(len(list_vals))])
            else:
                row.extend([None] * (1 + len(list_vals)))
        ws_c.append(row)

    wb.save(out_path)


def read_grid_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None
    header = list(rows[0])
    # SOUNDING | v1 | v2 ...
    start = 1 if header and (str(header[0]).upper() == "SOUNDING" or to_num(header[0]) is None) else 0
    vals = []
    for c in range(start, len(header)):
        n = to_num(header[c])
        if n is None:
            break
        vals.append(n)
    if len(vals) < 1:
        return None
    axis, grid = [], []
    for row in rows[1:]:
        if not row:
            continue
        s = to_num(row[0])
        if s is None:
            continue
        axis.append(s)
        grid.append([to_num(row[j + 1]) or 0 for j in range(len(vals))])
    if len(axis) < 1:
        return None
    return {"vals": vals, "axis": axis, "grid": grid}


def read_volume_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    x, v = [], []
    for i, row in enumerate(rows):
        if not row:
            continue
        a, b = to_num(row[0] if len(row) > 0 else None), to_num(row[1] if len(row) > 1 else None)
        if a is None or b is None:
            continue
        if i == 0 and str(rows[0][0]).upper() == "SOUNDING":
            continue
        x.append(a)
        v.append(b)
    if len(x) < 2:
        return None
    return {"x": x, "v": v}


def import_xlsx(path: str) -> dict:
    wb = load_workbook(path, data_only=True)
    patch = {}

    if "META" in wb.sheetnames:
        ws = wb["META"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            val = row[1] if len(row) > 1 else None
            if key == "id":
                continue
            if key not in META_KEYS:
                continue
            if key in ("calcType", "soundingMethod", "name"):
                patch[key] = "" if val is None else str(val)
            else:
                n = to_num(val)
                patch[key] = n if n is not None else val

    if "TRIM" in wb.sheetnames:
        g = read_grid_sheet(wb["TRIM"])
        if g and g["axis"]:
            patch["trimAxis"] = g["axis"]
            patch["trimVals"] = g["vals"]
            patch["trimGrid"] = g["grid"]
            patch.setdefault("soundingIncrement", detect_inc(g["axis"]))

    if "LIST" in wb.sheetnames:
        g = read_grid_sheet(wb["LIST"])
        if g and g["axis"]:
            patch["listAxis"] = g["axis"]
            patch["listVals"] = g["vals"]
            patch["listGrid"] = g["grid"]
            patch.setdefault("heelIncrement", detect_inc(g["axis"]))

    if "VOLUME" in wb.sheetnames:
        vol = read_volume_sheet(wb["VOLUME"])
        if vol:
            patch["volumeCurve"] = vol
            patch.setdefault("capacity", max(vol["v"]) if vol["v"] else 0)
            patch.setdefault("soundingIncrement", detect_inc(vol["x"]))

    # Fallback: single sheet that looks like a grid (first sheet / TABLE)
    if not patch.get("trimAxis") and not patch.get("volumeCurve"):
        for name in wb.sheetnames:
            g = read_grid_sheet(wb[name])
            if g and len(g["axis"]) >= 2 and len(g["vals"]) >= 2:
                patch["trimAxis"] = g["axis"]
                patch["trimVals"] = g["vals"]
                patch["trimGrid"] = g["grid"]
                patch.setdefault("soundingIncrement", detect_inc(g["axis"]))
                break

    if not patch.get("trimAxis") and not patch.get("listAxis") and not patch.get("volumeCurve"):
        raise ValueError("No TRIM/LIST/VOLUME tables found in workbook")
    return patch


def main():
    ap = argparse.ArgumentParser()
    sub = ap.add_subparsers(dest="cmd", required=True)

    ex = sub.add_parser("export")
    ex.add_argument("--in", dest="infile", required=True, help="Tank JSON file or - for stdin")
    ex.add_argument("--out", dest="outfile", required=True)

    im = sub.add_parser("import")
    im.add_argument("--in", dest="infile", required=True)

    args = ap.parse_args()
    try:
        if args.cmd == "export":
            if args.infile == "-":
                tank = json.load(sys.stdin)
            else:
                with open(args.infile, "r", encoding="utf-8") as f:
                    tank = json.load(f)
            export_xlsx(tank, args.outfile)
            json.dump({"ok": True, "out": args.outfile}, sys.stdout)
        else:
            patch = import_xlsx(args.infile)
            json.dump({"ok": True, "calibration": patch}, sys.stdout)
    except Exception as e:
        json.dump({"error": str(e)}, sys.stdout)
        sys.exit(1)


if __name__ == "__main__":
    main()
